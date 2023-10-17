import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl

# ... (Configura tus credenciales de correo y detalles del servidor SMTP)

# Crear un archivo Excel con la información del formulario
form_data = {"CUI": cui}  # Reemplaza cui con el valor del formulario
# Agrega más campos del formulario como sea necesario

workbook = openpyxl.Workbook()
sheet = workbook.active

for i, (field, value) in enumerate(form_data.items(), start=1):
    sheet.cell(row=i, column=1, value=field)
    sheet.cell(row=i, column=2, value=value)

workbook.save("formulario.xlsx")

# Adjuntar el archivo al correo
message = MIMEMultipart()
# ...

# Adjuntar el archivo al mensaje
with open("formulario.xlsx", "rb") as file:
    attachment = MIMEText(file.read(), 'base64')
    attachment.add_header('Content-Disposition', 'attachment', filename='formulario.xlsx')
    message.attach(attachment)

message.attach(MIMEText(body, 'plain'))

# Inicia la conexión SMTP y envía el correo
try:
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    server.login(from_email, password)
    server.sendmail(from_email, to_email, message.as_string())
    server.quit()
    print("Correo enviado con éxito")
except Exception as e:
    print(f"Error al enviar el correo: {str(e)}")
